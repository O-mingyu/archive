import datetime
import pandas as pd
import numpy as np
import copy
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


class Calculator:
    def __init__(self, production_file, order_file, order_change_file, cost_file):
        self.cost = pd.read_excel(cost_file)
        self.order = pd.read_excel(order_file)
        self.production = pd.read_excel(production_file)
        self.order_change = pd.read_excel(order_change_file)
        self.production['Date'] = pd.to_datetime(self.production['Date'])
        self.production['Date'] = self.production['Date'].apply(lambda x: x.replace(second=0, microsecond=0))
        self.result = pd.DataFrame(columns=['주문번호', '제품번호', '제품명', '생산량', '셋업량', '공정시작일', '공정마감일', '공정시작인덱스', '공정마감인덱스'])
        self.result['주문번호'] = self.order['주문번호']
        self.result['제품번호'] = self.order['제품번호']
        self.result['제품명'] = self.order['제품명']
        self.result['생산량'] = 0
        self.result['셋업량'] = 0
        self.result['공정시작일'] = 0
        self.result['공정마감일'] = 0
        self.result['공정시작인덱스'] = 0
        self.result['공정마감인덱스'] = 0
        self.result['주문수익'] = 0
        # self.result['주문공정시간'] = 0

    def disc_order(self, order_num, prod_table):
        target_order = self.order[self.order['주문번호'] == order_num]
        target_pro = target_order.iloc[0, 1]
        target_model = prod_table[target_pro]
        blk_list = []
        if len(target_model[target_model == 's']) == 96:
            start_idx = target_model[target_model == 's'].index[0]
            blk_list.append(start_idx)
            for i in range(start_idx+48, len(target_model)):
                if target_model.loc[i] == 0:
                    end_idx = i-1
                    blk_list.append(end_idx)
                    break
                elif target_model.loc[i] == 's':
                    end_idx = i - 1
                    blk_list.append(end_idx)
                    break
                elif i == len(target_model) - 1:
                    end_idx = i
                    blk_list.append(end_idx)
                    break

            start_idx2 = target_model[target_model == 's'].index[48]
            blk_list.append(start_idx2)
            for j in range(start_idx2, len(target_model)):
                if target_model.loc[j] == 0:
                    end_idx2 = j-1
                    blk_list.append(end_idx2)
                    break
                elif j == len(target_model) - 1:
                    end_idx2 = j
                    blk_list.append(end_idx2)
                    break
        elif len(target_model[target_model == 's']) == 48:
            start_idx = target_model[target_model == 's'].index[0]
            blk_list.append(start_idx)
            for i in range(start_idx, len(target_model)):
                if target_model.loc[i] == 0:
                    end_idx = i - 1
                    blk_list.append(end_idx)
                    break
                elif i == len(prod_table) - 1:
                    end_idx = i
                    blk_list.append(end_idx)
                    break
        elif len(target_model[target_model == 's']) >= 48 and len(target_model[target_model == 's']) < 96:
            start_idx = target_model[target_model == 's'].index[0]
            if target_model.loc[start_idx+47] != 's':
                start_idx = target_model[target_model == 's'].index[-1] - 47
            blk_list.append(start_idx)
            for i in range(start_idx + 48, len(target_model)):
                if target_model.loc[i] == 0:
                    end_idx = i - 1
                    blk_list.append(end_idx)
                    break
                elif target_model.loc[i] == 's':
                    end_idx = i - 1
                    blk_list.append(end_idx)
                    break
                elif i == len(prod_table) - 1:
                    end_idx = i
                    blk_list.append(end_idx)
                    break

        elif len(target_model[target_model == 's']) < 48 and len(target_model[target_model == 's']) > 0:
            start_idx = target_model[target_model == 's'].index[0]
            blk_list.append(start_idx)
            for i in range(start_idx, len(target_model)):
                if target_model.loc[i] == 0:
                    end_idx = i - 1
                    blk_list.append(end_idx)
                    break
                elif i == len(prod_table) - 1:
                    end_idx = i
                    blk_list.append(end_idx)
                    break
        else:
            start_idx = 0
            blk_list.append(start_idx)
            end_idx = 0
            blk_list.append(end_idx)
        first_numb = blk_list[0]
        last_numb = blk_list[1]
        same_order = self.order[self.order['제품명'] == target_order.iloc[0, 1]]
        same_order = same_order.reset_index()
        diff_order = same_order[same_order['주문번호'] == target_order['주문번호'].values[0]]
        index_order = diff_order.index
        for a in range(len(same_order)):
            if index_order[-1] == 1:
                first_numb = blk_list[-2]
                last_numb = blk_list[-1]
            elif index_order == a:
                first_numb = blk_list[0]
                last_numb = blk_list[1]
        # if order_num >= 10 and len(blk_list) == 4:
        #     first_numb = blk_list[2]
        #     last_numb = blk_list[3]
        return first_numb, last_numb, blk_list

    def get_production_quantity(self, order_num, prod_table):
        target_order = self.order[self.order['주문번호'] == order_num]
        target_pro = target_order.iloc[0, 1]
        idx = self.disc_order(order_num, prod_table)
        production_quantity = 0
        for i in range(idx[0], idx[1]+1):
            if prod_table.loc[i, target_pro] != 's':
                production_quantity += prod_table.loc[i, target_pro]
        return production_quantity

    def result_table(self, i, prod_table):
        target_order = self.order[self.order['주문번호'] == i]
        target_pro = target_order.iloc[0,1]
        result = self.get_production_quantity(i, prod_table)
        numb = self.disc_order(i, prod_table)
        start_date = prod_table.loc[numb[0], 'Date']
        due_date = prod_table.loc[numb[1], 'Date']
        setup_counts = prod_table.loc[numb[0]:numb[1]+1, target_pro].apply(lambda x: x == 's').sum().sum()
        self.result.loc[self.result['주문번호'] == i, '생산량'] = result
        self.result.loc[self.result['주문번호'] == i, '공정시작일'] = start_date
        self.result.loc[self.result['주문번호'] == i, '공정마감일'] = due_date
        self.result.loc[self.result['주문번호'] == i, '공정시작인덱스'] = numb[0]
        self.result.loc[self.result['주문번호'] == i, '공정마감인덱스'] = numb[1]
        self.result.loc[self.result['주문번호'] == i, '셋업량'] = setup_counts

    def c_profit(self):
        for i in range(len(self.order)):
            target_order = self.order[self.order['주문번호'] == i]
            target_num = target_order.iloc[0, 2]
            self.result.loc[self.result['주문번호'] == i, '주문수익'] = (self.result.loc[self.result['주문번호'] == i, '생산량'].iloc[0] * (
                        self.cost.loc[target_num, '이익'] - self.cost.loc[target_num, '생산비용']) - self.result['셋업량'] *
                                                                 self.cost.loc[target_num, '셋업비용'])
            # self.result.loc[self.result['주문번호'] == i, '주문공정시간'] = self.result['셋업량'] + self.result.loc[
            #     self.result['주문번호'] == i, '생산량'] * self.cost.loc[target_num, '생산시간']


    def calc_profit(self, order_num):
        return self.result['주문수익'].values[order_num]

    def c_time(self, order_table):  # 생산계획에서 생산량 셋업량 계산 후 생산시간 반환
        for i, _ in order_table.iterrows():
            target_order = order_table[order_table['주문번호'] == 0]
            target_num = target_order.iloc[0, 2]
            self.result.loc[self.result['주문번호'] == i, '주문공정시간'] = self.result['셋업량'] + self.result.loc[self.result['주문번호'] == i, '생산량'] * self.cost.loc[target_num, '생산시간']
        
    def calc_time(self, order_num):  # 제품 넘버에 해당하는 제품생산시간 반환 c_time 뒤에 와야함
        return self.result['주문공정시간'].values[order_num]

    def time_check(self, order_num, prod_table, order_table):
        prod_table['Date'] = pd.to_datetime(prod_table['Date'])
        order_table['납기일'] = pd.to_datetime(order_table['납기일'])
        target_order = order_table[order_table['주문번호'] == order_num]
        due_date = pd.to_datetime(target_order.iloc[0, 4])
        end_date = pd.to_datetime(self.result.loc[self.order['주문번호'] == order_num, '공정마감일'].iloc[0])
        end_idx = self.result.loc[self.order['주문번호'] == order_num, '공정마감인덱스'].iloc[0]
        if end_date == due_date:  # 납기일과 마감일 동일
            return 0
        elif end_date < due_date and end_idx != 0:  # 마감일이 납기일보다 빨랐을 경우
            return 1
        elif end_date > due_date and end_date <= due_date + pd.Timedelta(days=10):  # 납기일을 못지켰지만 납기일 +10일까지 마감일이 들어올 때
            return 2
        elif end_date > due_date + pd.Timedelta(days=10):  # 마감일이 납기일 10일을 넘기는 경우
            return 3
        elif end_idx == 0:
            return 4

    def table_time(self, prod_table, order_table):
        for order_num, _ in order_table.iterrows():
            x = self.time_check(order_num, prod_table, order_table)
            if x == 3:
                return 3
            else:
                pass
        return 0

    def table_check(self, order_num, prod_table, order_table):
        target_order = order_table[order_table['주문번호'] == order_num]
        target_num = target_order.iloc[0, 2]
        due_date = pd.to_datetime(target_order.iloc[0, 4])
        due_date = due_date.replace(second=0, microsecond=0)
        end_date = pd.to_datetime(self.result.loc[order_num, '공정마감일'])
        end_date = end_date.replace(second=0, microsecond=0)
        x = self.time_check(order_num, prod_table, order_table)
        if x == 0:
            self.result.loc[order_num, '주문수익'] = self.calc_profit(order_num)
        elif x == 1:
            gap = ((due_date - end_date).total_seconds() / 3600) // 24
            self.result.loc[order_num, '주문수익'] = self.calc_profit(order_num) - gap * self.cost.loc[
                target_num, '재고보유비용']
        elif x == 2:
            gap = ((end_date - due_date).total_seconds() / 3600) // 24
            self.result.loc[order_num, '주문수익'] = self.calc_profit(order_num) - gap * (self.cost.loc[
                target_num, '납기 지연 패널티 비용'])
        elif x == 3:
            gap = ((end_date - due_date).total_seconds() / 3600) // 24
            self.result.loc[order_num, '주문수익'] = self.calc_profit(order_num) - gap * (
            self.cost.loc[target_num, '납기 지연 패널티 비용']) - (order_table.loc[order_num, '주문량'] - self.result.loc[
                order_num, '생산량']) * self.cost.loc[target_num, '납품 미달 패널티 비용']
        elif x == 4:
            self.result.loc[order_num, '주문수익'] = 0 - order_table.loc[order_num, '주문량'] * self.cost.loc[target_num, '납품 미달 패널티 비용']
        return


    def new_table(self):  # 기존 파일과 새로운 파일의 변동된 부분 출력
        merge_table = self.order.merge(self.order_change, indicator=True, how='outer')
        diff = merge_table.loc[merge_table['_merge'] == 'right_only'].reset_index(drop=True)
        diff['주문 변화량'] = self.order['주문량'] - self.order_change['주문량']
        diff['납기 변화량'] = self.order['납기일'] - self.order_change['납기일']
        return diff

    def order_idx(self):
        idx_list = []
        for i in range(len(self.order)):
            x = self.disc_order(i,self.production)
            idx_list.append(x[:2])
        return idx_list

    def bruteforce(self, order_num):
        new_order = self.new_table()
        if order_num not in new_order['주문번호'].unique():
            return print('해당 주문은 변동이 없습니다')
        table_profit = []
        result_plan = self.production
        product_num = self.order_change.loc[self.order_change['주문번호'] == order_num, '제품번호'].iloc[0]
        target_pro = self.order_change.loc[self.order_change['주문번호'] == order_num, '제품명'].iloc[0]
        same_order = self.order_change[self.order_change['제품명'] == target_pro]
        target_duedate = self.order_change.loc[self.order_change['주문번호'] == order_num, '납기일'].iloc[0]
        target_duedate = pd.to_datetime(target_duedate)
        duedate_idx = self.production[self.production['Date'] == target_duedate].index[0]
        profit_list = []
        order_amount = self.order_change.loc[order_num, '주문량']
        hourly_production = 1 / (self.cost['생산시간'].values[product_num])
        value = order_amount / hourly_production
        order_idx = self.order_idx()
        p_idx = self.disc_order(order_num, self.production)
        for index in range(0, self.production.index[-1]):
            print(target_pro, index)
            new_plan = copy.deepcopy(self.production)
            new_plan.loc[p_idx[0]:p_idx[1], target_pro] = 0
            setup_s = index
            setup_e = index + 47
            prod_s = index + 48
            prod_e = index + 47 + int(value)
            if prod_e > len(new_plan):
                break
            if prod_e > duedate_idx + 10:
                prod_e = duedate_idx + 10
            if prod_s >= prod_e:
                break
            for tup in order_idx:
                if index in range(tup[0], tup[1]+1) or prod_e in range(tup[0], tup[1]+1):
                    nontarget_num = order_idx.index(tup)
                    nontarget_pro = self.order_change.loc[self.order_change['주문번호'] == nontarget_num, '제품명'].iloc[0]
                    new_plan.loc[tup[0]:tup[1]+1, nontarget_pro] = 0
                    print('겹치는 생산계획', nontarget_pro)
            new_plan.loc[setup_s: setup_e, target_pro] = 's'
            new_plan.loc[prod_s: prod_e, target_pro] = hourly_production
            new_plan.loc[setup_s: prod_e, new_plan.columns.difference(['Date', target_pro])] = 0
            for ii in range(len(self.order_change)):  # 생성된 계획의 생산량, 공정마감일 계산
                self.result_table(ii, new_plan)
            if order_num < 10:
                self.result.loc[self.result.duplicated(subset=['공정마감인덱스'], keep='first'), ['공정마감인덱스', '공정마감일', '생산량']] = 0
            elif order_num >= 10:
                self.result.loc[
                    self.result.duplicated(subset=['공정마감인덱스'], keep='last'), ['공정마감인덱스', '공정마감일', '생산량']] = 0
            temp_idx = self.disc_order(order_num, new_plan)
            print(temp_idx)
            if order_num >= 10 and index < p_idx[2][0]:  # 주문2의 탐색이 탐색하면서 첫번째 주문보다 앞을 탐색하게 되는 경우
                print('주문2의 탐색 범위가 주문1보다 앞입니다.')
                self.result.loc[self.result['주문번호'] == order_num].iloc[0, 3:], self.result.loc[self.result['주문번호'] == same_order.iloc[0,0]].iloc[0, 3:] = self.result.loc[self.result['주문번호'] == same_order.iloc[0,0]].iloc[0, 3:], self.result.loc[self.result['주문번호'] == order_num].iloc[0, 3:]
                # self.result.loc[self.result['주문번호'] == order_num, '공정마감인덱스'].iloc[0] = temp_idx[2][1]  # blk_list==[두번째주문시작, 두번째주문마감, 첫번째주문시작, 첫번째주문마감] 이 되어 두번째 주문마감을 선택
                # p_quantity = 0
                # for t in range(temp_idx[2][0], temp_idx[2][1] + 1):
                #     if new_plan.loc[t, target_pro] != 's':
                #         p_quantity += new_plan.loc[t, target_pro]
                # self.result.loc[self.result['주문번호'] == order_num, '생산량'].iloc[0] = p_quantity
            self.c_profit()
            for jj in range(len(self.order_change)):
                self.table_check(jj, new_plan, self.order)
            self.table_check(order_num, new_plan, self.order_change)
            x = sum(self.result['주문수익'])
            profit_list.append(x)
        max_int = max(profit_list)
        result_plan.loc[p_idx[0]:p_idx[1], target_pro] = 0
        max_idx = profit_list.index(max_int)
        for tup in order_idx:
            if max_idx in range(tup[0], tup[1] + 1) or max_idx + int(value) + 47 in range(tup[0], tup[1] + 1):
                nontarget_num = order_idx.index(tup)
                nontarget_pro = self.order_change.loc[self.order_change['주문번호'] == nontarget_num, '제품명'].iloc[0]
                result_plan.loc[tup[0]:tup[1] + 1, nontarget_pro] = 0
        if max_idx + int(value) + 47 > duedate_idx + 10:
            result_plan.loc[max_idx:max_idx + 47, target_pro] = 's'
            result_plan.loc[max_idx + 48:duedate_idx + 10, target_pro] = hourly_production
            result_plan.loc[max_idx:duedate_idx + 10, result_plan.columns.difference(['Date', target_pro])] = 0
        result_plan.loc[max_idx:max_idx + 47, target_pro] = 's'
        result_plan.loc[max_idx + 48:max_idx + 47 + int(value), target_pro] = hourly_production
        result_plan.loc[max_idx:max_idx + 47 + int(value), result_plan.columns.difference(['Date', target_pro])] = 0
        for kk in range(len(self.order)):
            self.result_table(kk, result_plan)
        if order_num < 10:
            self.result.loc[self.result.duplicated(subset=['공정마감인덱스'], keep='first'), ['공정마감인덱스', '공정마감일', '생산량']] = 0
        else:
            self.result.loc[self.result.duplicated(subset=['공정마감인덱스'], keep='last'), ['공정마감인덱스', '공정마감일', '생산량']] = 0
        if order_num >= 10 and max_idx < p_idx[2][0]:  # 주문2의 경우에서 최적해가 주문1의 인덱스보다 앞일 때
            w_idx = self.disc_order(order_num, result_plan)
            print('주문2가 주문1보다 먼저입니다')
            self.result.loc[self.result['주문번호'] == order_num].iloc[0, 3:], self.result.loc[
                                                                               self.result['주문번호'] == same_order.iloc[
                                                                                   0, 0]].iloc[0, 3:] = self.result.loc[
                                                                                                            self.result[
                                                                                                                '주문번호'] ==
                                                                                                            same_order.iloc[
                                                                                                                0, 0]].iloc[
                                                                                                        0, 3:], \
            self.result.loc[self.result['주문번호'] == order_num].iloc[0, 3:]
        self.c_profit()
        for kk in range(len(self.order_change)):
            self.table_check(kk, result_plan, self.order)
        self.table_check(order_num, result_plan, self.order_change)
        now = datetime.datetime.now()
        filename1 = 'C:/Users/Admin/Desktop/GPS/계산기/최적_계획_{0}_{1}.xlsx'.format(order_num, now.strftime('%Y%m%d_%H%M%S'))
        filename2 = 'C:/Users/Admin/Desktop/GPS/계산기/공정결과표_{0}_{1}.xlsx'.format(order_num, now.strftime('%Y%m%d_%H%M%S'))
        with pd.ExcelWriter(filename1, engine='openpyxl') as writer:
            result_plan.to_excel(writer, index=False)
        f_profit = sum(self.result['주문수익'])
        table_profit.append(f_profit)
        with pd.ExcelWriter(filename2, engine='openpyxl') as writer:
            self.result.to_excel(writer, index=False)

        # original_table = load_workbook(filename='생산_시간별_시퀀스.xlsx')
        # new_table = load_workbook(filename='최적_계획.xlsx')
        # yellow_fill = PatternFill(start_color="FFFF09", end_color="FFFF09", fill_type="solid")
        # sheet1 = original_table.active
        # sheet2 = new_table.active
        #
        # for row in range(1, sheet2.max_row + 1):
        #     for col in range(2, sheet2.max_column + 1):
        #         cell1 = sheet1.cell(row=row, column=col)
        #         cell2 = sheet2.cell(row=row, column=col)
        #         if cell1.value != cell2.value:
        #             cell2.fill = yellow_fill
        # new_table.save('최적_계획.xlsx')
        return print(profit_list, max_int, self.result)



Calc = Calculator('생산_시간별_시퀀스.xlsx', '주문_테이블.xlsx', '주문_변동_테이블_6_4.xlsx', '비용_테이블.xlsx')

production_table = pd.read_excel('생산_시간별_시퀀스.xlsx')
production_table['Date'] = pd.to_datetime(production_table['Date'])
production_table['Date'] = production_table['Date'].apply(lambda x: x.replace(second=0, microsecond=0))
order_table = pd.read_excel('주문_테이블.xlsx')
order_change_table = pd.read_excel('주문_변동_테이블_6_4.xlsx')
cost_table = pd.read_excel('비용_테이블.xlsx')
start = time.time()
for k in range(len(order_change_table)):
    Calc.bruteforce(k)

print(time.time() - start)
for i in range(14):
    Calc.result_table(i, production_table)
Calc.c_profit()
for j in range(14):
    Calc.table_check(j, production_table, order_change_table)
print(Calc.result, sum(Calc.result['주문수익']))
