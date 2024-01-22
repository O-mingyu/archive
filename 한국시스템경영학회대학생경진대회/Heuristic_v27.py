import pandas as pd
import math
import os
import datetime
import copy
import time
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

class Heuristic_Production_Plan:
    def __init__(self, production_plan, cost_table, order_table, order_change_table):
        self.production_plan = pd.read_excel(production_plan)
        self.cost_table = pd.read_excel(cost_table)
        self.order_table = pd.read_excel(order_table)
        self.order_change_table = pd.read_excel(order_change_table)
        self.production_plan['Date']=pd.to_datetime(self.production_plan['Date'])
        self.production_plan['Date']=self.production_plan['Date'].apply(lambda x: x.replace(second=0, microsecond=0))
        self.result = pd.DataFrame(columns=['주문번호', '제품번호', '제품명', '생산량', '셋업량', '공정마감일', '공정시작일', '공정시작인덱스','공정마감인덱스'])
        self.result['주문번호'] = self.order_table['주문번호']
        self.result['제품번호'] = self.order_table['제품번호']
        self.result['제품명'] = self.order_table['제품명']
        self.result['생산량'] = 0
        self.result['셋업량'] = 0
        self.result['공정마감일'] = 0
        self.result['공정시작일'] = 0
        self.result['공정시작인덱스'] = 0
        self.result['공정마감인덱스'] = 0
        self.result['주문수익'] = 0
        self.result['공정시간'] = 0
        
    def determine(self, order_number): #주문번호를 받아 해당 주문이 증산인지, 감산인지, 납기일 변동인지 파악
        combine_table = pd.merge(self.order_table, self.order_change_table, on='주문번호', suffixes=('_1', '_2'))
        
        #주문량차이, 납기일차이 계산 
        combine_table['주문량차이'] = combine_table['주문량_2'] - combine_table['주문량_1']
        combine_table['납기일차이'] = (pd.to_datetime(combine_table['납기일_2']) - pd.to_datetime(combine_table['납기일_1'])).dt.days
        
        differ_table = combine_table.loc[(combine_table['주문량차이'] != 0) | (combine_table['납기일차이'] != 0)]
        result_table = differ_table[['주문번호', '제품명_1', '제품번호_1',  '주문량차이',  '납기일차이']]
        
        # 변동 주문번호인지 확인
        if order_number not in result_table['주문번호']: 
            print("변동 주문이 들어오지 않은 주문번호 입니다.")
            return None, None
        
        target = result_table[result_table['주문번호'] == order_number]
        
        if target['주문량차이'].values[0] > 0:
            operation = 'increment'
        elif target['주문량차이'].values[0] < 0:
            operation = 'decrement'
        elif target['납기일차이'].values[0] != 0:
            operation = 'deadline'
            
        return target, operation

    def make_blk(self, order_number, pro_plan):
        target_order = self.order_table[self.order_table['주문번호'] == order_number]
        target_pro = target_order.iloc[0, 1]  # 제품명
        same_order = self.order_table[self.order_table['제품명'] == target_pro]
        # 우리가 만들고자 하는 주문의 생산계획 정보
        target_model = pro_plan[[target_pro]]

        # 우리가 만들고자 하는 주문이 가지는 블럭 추출
        blk_list = []  # 엑셀 상의 행 번호와 2만큼 차이남
        j = 0
        for _ in range(len(same_order)):
            for k in range(j, len(pro_plan)):
                if target_model.iloc[k, 0] != 0:
                    blk_list.append(k)
                    break
            for r in range(k, len(pro_plan)):
                # if r + 1 == len(pro_plan):
                #     blk_list.append(r)
                #     break
                if target_model.iloc[r, 0] == 0:
                    blk_list.append(r - 1)
                    break
                elif len(blk_list) >= 3 and r + 1 == len(pro_plan):
                    blk_list.append(r)
                    break
            j = r
        # blk_list에서 제품번호에 맞는 인덱스 추출
        if len(same_order) == 1:
            pass
        elif len(blk_list) == 4:
            if order_number == same_order.iloc[0][0]:  # 주문 번호가 맨 앞인 주문
                del blk_list[2:]
            if order_number == same_order.iloc[1][0]:  # 주문 번호가 두번째이자 마지막인 주문
                del blk_list[:2]
                del blk_list[2:]
            # 동일 제품 증산 주문이 3개 이상일때
        if len(same_order) >= 3:
            for i in range(1, len(same_order) - 1):
                if order_number == same_order.iloc[i][0]:
                    del blk_list[:2]
                    del blk_list[2:]
        first_num = blk_list[0]
        last_num = blk_list[1]
        return first_num, last_num, blk_list

    def increment_plan(self, target, order_number): #증산 함수
        #필요 변수 셋팅
        target_order = self.order_change_table[self.order_change_table['제품번호']==target['제품번호_1'].values[0]] #  주문 정보
        target_cost = self.cost_table[self.cost_table['제품번호'] == target['제품번호_1'].values[0]]  # 해당 주문의 비용 정보
        target_num = target_order.iloc[0, 2]  # 제품 번호
        target_pro = target_order.iloc[0, 1]  # 제품명
        dif_amount = target.iloc[0,3]  # 증산량
        dif_profit = target_cost.iloc[0,8]  # 생산 이익
        dif_cost = target_cost.iloc[0,6]  # 생산 비용
        dif_setup = target_cost.iloc[0, 7] # 셋업 비용
        per_produce = int(1 / target_cost.iloc[0,2])  # 생산 가능량
        require_cell = math.ceil(dif_amount / per_produce)  # 생산하는데 필요한 셀 개수(올림)
        pro_plan = self.production_plan
        due_date = pd.to_datetime(target_order.iloc[0, 4])
        due_date = due_date.replace(second=0, microsecond=0) # 납기일
        target_model = pro_plan.loc[:, ['Date', target_pro]] # 주문 번호 해당 제품의 생산계획 열 추출
        first_num, last_num, blk_list = self.make_blk(order_number, pro_plan) # 주문 번호 해당 공정의 인덱스 추출
        target_profit_0 = dif_amount * (dif_profit - dif_cost) - 48 * dif_setup # 증산량 수익(이익 - 생산비 - 셋업비)
        target_area = target_model.iloc[first_num:last_num + 1 + require_cell]  # 증산을 뒤에 붙이는 경우 필요한 영역
        using_cell = last_num-(first_num+48) # 현재 생산에 사용되는 셀 개수
        compare_date = pd.to_datetime(target_area.iloc[-1, 0]) #증산을 뒤에 붙이는 경우 필요한 영역의 마지막 일자
        idx_list = []
        for ii in range(len(self.order_table)):  # 생산계획 테이블 상 공정들의 인덱스 추출
            numb = self.make_blk(ii, pro_plan)
            idx_list.append([numb[0], numb[1]])

        # changed_start_idx_1 = first_num - (require_cell + 48)
        # changed_end_idx_1 = changed_start_idx_1 + 48 + require_cell + using_cell
        # changed_start_idx_2 = first_num
        # changed_end_idx_2 = last_num + require_cell

        #증산 계산
        if compare_date <= due_date + pd.Timedelta(days=10): # 뒤로 붙이는 경우
            new_last_num = last_num + require_cell
            end_date = pd.to_datetime(pro_plan['Date'][new_last_num])
            if end_date <= due_date: # 마감일이 납기일 보다 빠를 때
                gap = ((due_date - end_date).total_seconds() / 3600) // 24
                target_profit = target_profit_0 - gap * self.cost_table.loc[target_num, '재고보유비용']
            elif end_date > due_date and end_date <= due_date + pd.Timedelta(days=10): # 마감일이 납기일+10 이내일 때
                gap = ((end_date - due_date).total_seconds() / 3600) // 24
                target_profit = target_profit_0 - gap * self.cost_table.loc[target_num, '납기 지연 패널티 비용']
            non_target = []
            for tup in idx_list:  # 변동한 공정과 겹치는 공정 파악
                if first_num in range(tup[0], tup[1] + 1) or new_last_num in range(tup[0], tup[1] + 1):
                    nontarget_num = idx_list.index(tup)
                    nontarget_pro = \
                    self.order_change_table.loc[self.order_change_table['주문번호'] == nontarget_num, '제품명'].iloc[0]
                    non_target.append([nontarget_num, nontarget_pro])
            total_profit = 0
            if non_target is not None:  # 겹치는 공정들의 수익 + 납품하지 않았을 시 패널티 비용 계산
                for jj in non_target:
                    if jj[0] == order_number:
                        continue
                    nontarget_result = self.get_production_quantity(jj[0])
                    total_profit += nontarget_result * (
                                self.cost_table.loc[self.cost_table['제품명'] == jj[1], '이익'].iloc[0] + self.cost_table.loc[
                            self.cost_table['제품명'] == jj[1], '납품 미달 패널티 비용'].iloc[0] - self.cost_table.loc[
                                    self.cost_table['제품명'] == jj[1], '생산비용'].iloc[0]) - 48 * self.cost_table.loc[
                                        self.cost_table['제품명'] == jj[1], '셋업비용'].iloc[0]
            #뒤에다가 붙임
            if target_profit >= total_profit: # 변동 주문의 수익이 겹치는 주문의 수익보다 클 때
                print(f'{target_pro} 증산: 뒤에 붙임 변동 주문 수익: {target_profit}, 겹치는 공정 수익: {total_profit}')
                if non_target is not None:
                    for ll in non_target:  # 겹치는 공정 삭제
                        if ll[0] == order_number:
                            continue
                        nontarget_idx = self.make_blk(ll[0], pro_plan)
                        print(ll[0], nontarget_idx)
                        pro_plan.loc[nontarget_idx[0] : nontarget_idx[1] + 1, ll[1]] = 0
                pro_plan.loc[last_num : new_last_num, target_pro] = per_produce
            else: # 크지 않다면 앞으로 붙일 수 있는지 확인
                new_first_num = first_num - (require_cell + 48)
                new_last_num = new_first_num + 48 + require_cell + using_cell
                end_date = pd.to_datetime(pro_plan['Date'][new_last_num])
                gap = (((due_date - end_date).total_seconds()) / 3600) // 24
                target_profit = target_profit_0 - gap * self.cost_table.loc[
                    target_num, '재고보유비용']  # 생산량 * (이익 - 생산비용) - 셋업량 * 셋업비용 - (납기일 - 마감일) * 재고보유패널티
                idx_list = []
                for ii in range(len(self.order_table)):  # 생산계획 테이블 상 공정들의 인덱스 추출
                    numb = self.make_blk(ii, pro_plan)
                    idx_list.append([numb[0], numb[1]])
                non_target = []
                for tup in idx_list:  # 변동한 공정과 겹치는 공정 파악
                    if new_first_num in range(tup[0], tup[1] + 1) or new_last_num in range(tup[0], tup[1] + 1):
                        nontarget_num = idx_list.index(tup)
                        nontarget_pro = \
                            self.order_change_table.loc[
                                self.order_change_table['주문번호'] == nontarget_num, '제품명'].iloc[0]
                        non_target.append([nontarget_num, nontarget_pro])
                total_profit = 0
                if non_target is not None:  # 겹치는 공정들의 수익 + 납품하지 않았을 시 패널티 비용 계산
                    for jj in non_target:
                        if jj[0] == order_number:
                            continue
                        nontarget_result = self.get_production_quantity(jj[0])
                        total_profit += nontarget_result * (
                                self.cost_table.loc[self.cost_table['제품명'] == jj[1], '이익'].iloc[0] +
                                self.cost_table.loc[
                                    self.cost_table['제품명'] == jj[1], '납품 미달 패널티 비용'].iloc[0] - self.cost_table.loc[
                                    self.cost_table['제품명'] == jj[1], '생산비용'].iloc[0]) - 48 * self.cost_table.loc[
                                            self.cost_table['제품명'] == jj[1], '셋업비용'].iloc[0]
                if target_profit >= total_profit:
                    print(f'{target_pro} 증산: 앞에다 붙임 변동 주문 수익: {target_profit}, 겹치는 공정 수익: {total_profit}')
                    if non_target is not None:
                        for ll in non_target:  # 겹치는 공정 삭제
                            if ll[0] == order_number:
                                continue
                            nontarget_idx = self.make_blk(ll[0], pro_plan)
                            print(ll[0], nontarget_idx)
                            pro_plan.loc[nontarget_idx[0]: nontarget_idx[1] + 1, ll[1]] = 0
                    pro_plan.loc[first_num: last_num, target_pro] = 0
                    pro_plan.loc[new_first_num: new_first_num + 47, target_pro] = 's'
                    pro_plan.loc[new_first_num + 48: new_last_num, target_pro] = per_produce
                else:
                    print(f'{target_pro} 증산: 미반영', target_profit, "<", total_profit)

        elif compare_date > due_date + pd.Timedelta(days=10): # 앞에 붙이는 경우
            new_first_num = first_num - (require_cell + 48)
            new_last_num = new_first_num + 48 + require_cell + using_cell
            end_date = pd.to_datetime(pro_plan['Date'][new_last_num])
            if end_date > due_date:
                print(f'{target_pro} 변동x 납기일 지킬 수 없음')
                return pro_plan
            else:
                gap = (((due_date - end_date).total_seconds()) / 3600) // 24
                target_profit = target_profit_0 - gap * self.cost_table.loc[
                    target_num, '재고보유비용'] # 생산량 * (이익 - 생산비용) - 셋업량 * 셋업비용 - (납기일 - 마감일) * 재고보유패널티
                non_target = []
                for tup in idx_list: # 변동한 공정과 겹치는 공정 파악
                    if new_first_num in range(tup[0], tup[1] + 1) or new_last_num in range(tup[0], tup[1] + 1):
                        nontarget_num = idx_list.index(tup)
                        nontarget_pro = self.order_change_table.loc[self.order_change_table['주문번호'] == nontarget_num, '제품명'].iloc[0]
                        non_target.append([nontarget_num, nontarget_pro])
                total_profit = 0
                if non_target is not None: # 겹치는 공정들의 수익 + 납품하지 않았을 시 패널티 비용 계산
                    for jj in non_target:
                        if jj[0] == order_number:
                            continue
                        nontarget_result = self.get_production_quantity(jj[0])
                        total_profit += nontarget_result * (
                                self.cost_table.loc[self.cost_table['제품명'] == jj[1], '이익'].iloc[0] +
                                self.cost_table.loc[
                                    self.cost_table['제품명'] == jj[1], '납품 미달 패널티 비용'].iloc[0] - self.cost_table.loc[
                                    self.cost_table['제품명'] == jj[1], '생산비용'].iloc[0]) - 48 * self.cost_table.loc[
                                            self.cost_table['제품명'] == jj[1], '셋업비용'].iloc[0]
                if target_profit >= total_profit:
                    print(f'{target_pro} 증산: 앞에다 붙임 변동 주문 수익: {target_profit}, 겹치는 공정 수익: {total_profit}')
                    if non_target is not None:
                        for ll in non_target: # 겹치는 공정 삭제
                            if ll[0] == order_number:
                                continue
                            nontarget_idx = self.make_blk(ll[0], pro_plan)
                            print(ll[0], nontarget_idx)
                            pro_plan.loc[nontarget_idx[0] : nontarget_idx[1] + 1, ll[1]] = 0
                    pro_plan.loc[first_num: last_num, target_pro] = 0
                    pro_plan.loc[new_first_num: new_first_num + 47, target_pro] = 's'
                    pro_plan.loc[new_first_num + 48: new_last_num, target_pro] = per_produce
                else:
                    print(f'{target_pro} 증산: 미반영', target_profit, "<", total_profit)
        return pro_plan

    def decrement_plan(self, target, order_number):  # 감산 함수
        target_order = self.order_change_table[self.order_change_table['제품번호'] == target['제품번호_1'].values[0]] # 변동 주문 정보
        target_cost = self.cost_table[self.cost_table['제품번호'] == target['제품번호_1'].values[0]]   # 변동 주문 가격 정보
        dif_amount = target.iloc[0, 3] # 감산량
        per_produce = int(1 / target_cost['생산시간'].values[0]) # 시간당 생산량
        require_cell = math.ceil(abs(dif_amount) / per_produce) # 감산시 줄여야 하는 셀의 개수
        pro_plan = self.production_plan
        target_pro = target_order.iloc[0, 1]  # 제품명
        target_model = pro_plan[[target_pro]] # 변동 제품 생산계획
        if (target_model == 0).all().values[0]:
            return pro_plan

        first_num, last_num, blk_list = self.make_blk(order_number, pro_plan)
        production_index = [first_num, last_num]
        
        #감산 계획 산출
        if len(production_index) == 0:
            print("감산 계획이 없습니다")
            return pro_plan
        elif len(production_index) == 2:
            pro_plan.loc[last_num - require_cell + 1: last_num, target_pro] = 0
            print(f'{target_pro} 감산')
            return pro_plan
                
    def deadline_plan(self, target, order_number): #납기일 변동 함수
        #우리가 현재 변동을 주고자 하는 주문에 대한 정보들
        target_order_0 = self.order_table[self.order_table['주문번호'] == target['주문번호'].values[0]]
        target_order = self.order_change_table[self.order_change_table['주문번호'] == target['주문번호'].values[0]]
        target_cost = self.cost_table[self.cost_table['제품번호'] == target['제품번호_1'].values[0]]
        target_num = target_order.iloc[0, 2]
        target_pro = target_order.iloc[0, 1]

        #납기일 변동이 가지는 주문을 만들기 위해 필요한 엑셀의 셀 수
        per_produce = (1 / target_cost['생산시간'].values[0]) 
        require_cell = math.ceil(target_order['주문량'].values[0] / per_produce) + 48

        #납기일 변동이 가지는 주문의 이익
        target_profit_0 = target_order['주문량'].values[0] * (target_cost['이익'].values[0] - target_cost['생산비용'].values[0]) - (target_cost['셋업비용'].values[0] * 48)

        #납기일을 맞추기 위해 필요로 하는 인덱스 넘버 추출
        pro_plan = copy.deepcopy(self.production_plan)
        m_row = self.make_blk(order_number, pro_plan)
        m_row_index = m_row[1] # 기존 공정 마감 인덱스
        delivery_date = pro_plan[pro_plan['Date'] == target_order.at[target_order['주문번호'].values[0],'납기일']]
        delivery_date_index = delivery_date.index[0] # 변동된 납기 인덱스
        result_plan = self.production_plan

        idx_list = []
        for ii in range(len(self.order_table)):  # 생산계획 테이블 상 공정들의 인덱스 추출
            numb = self.make_blk(ii, pro_plan)
            idx_list.append([numb[0], numb[1]])

        for p in range(0, 241, 24):
        # 생산에 시작과 끝의 인덱스를 24씩 더해주며 1일씩 연장
            pro_plan.loc[:,target_pro] = self.production_plan.loc[:, target_pro]
            last_num = delivery_date_index + p
            first_num = last_num - require_cell + 1
            gap = last_num - delivery_date_index
            target_profit = target_profit_0 - gap * self.cost_table.loc[target_num, '납기 지연 패널티 비용']
            non_target = []
            for tup in idx_list:  # 변동한 공정과 겹치는 공정 파악
                if first_num in range(tup[0], tup[1] + 1) or last_num in range(tup[0], tup[1] + 1):
                    nontarget_num = idx_list.index(tup)
                    nontarget_pro = \
                        self.order_change_table.loc[self.order_change_table['주문번호'] == nontarget_num, '제품명'].iloc[0]
                    non_target.append([nontarget_num, nontarget_pro])
            print(non_target)
            total_profit = 0
            if non_target is not None:  # 겹치는 공정들의 수익 + 납품하지 않았을 시 패널티 비용 계산
                for jj in non_target:
                    if jj[0] == order_number:
                        continue
                    nontarget_result = self.get_production_quantity(jj[0])
                    total_profit += nontarget_result * (
                            self.cost_table.loc[self.cost_table['제품명'] == jj[1], '이익'].iloc[0] + self.cost_table.loc[
                        self.cost_table['제품명'] == jj[1], '납품 미달 패널티 비용'].iloc[0] - self.cost_table.loc[
                                self.cost_table['제품명'] == jj[1], '생산비용'].iloc[0]) - 48 * self.cost_table.loc[
                                        self.cost_table['제품명'] == jj[1], '셋업비용'].iloc[0]
            if target_profit >= total_profit:
                for ll in non_target:  # 겹치는 공정 삭제
                    nontarget_idx = self.make_blk(ll[0], pro_plan)
                    print(ll[0], nontarget_idx)
                    non_target_pro = ll[1]
                    pro_plan.loc[nontarget_idx[0] : nontarget_idx[1] + 1, ll[1]] = 0
                pro_plan.loc[m_row[0]: m_row[1], target_pro] = 0
                pro_plan.loc[first_num:first_num + 47, target_pro] = 's'
                pro_plan.loc[first_num + 48: last_num, target_pro] = per_produce
                break
            else:
                continue
        # result_plan.loc[m_row[0]: m_row[1], target_pro] = 0
        # result_plan.loc[nontarget_idx[0]: nontarget_idx[1], non_target_pro] = 0
        # result_plan.loc[first_num: first_num + 47, target_pro] = 's'
        # result_plan.loc[first_num + 48: last_num, target_pro] = per_produce
        return pro_plan

    def make_new_table(self, order_number): #주문 번호에 따른 변동을 반영한 생산계획 생성
        target, operation = self.determine(order_number)
        if operation == 'increment':
            print(target['제품명_1'], '증산')
            final = self.increment_plan(target, order_number)
        elif operation == 'decrement':
            print(target['제품명_1'], '감산')
            final = self.decrement_plan(target, order_number)
        else:
            print(target['제품명_1'], '납기일 변동')
            final = self.deadline_plan(target, order_number)

        self.table_check()
        self.time_check()
        t_profit = self.profit(order_number)
        a_profit = self.all_profit()
        print(f"변동 주문 수익: {t_profit}, 전체 계획 수익: {a_profit}")
        # if os.path.isfile("C:/Users/82105/OneDrive/바탕 화면/계산기 데이터/new_생산계획.xlsx"):
        #     os.remove("C:/Users/82105/OneDrive/바탕 화면/계산기 데이터/new_생산계획.xlsx")      
        return final.to_excel("new_생산계획_test.xlsx", index=False)
    
    def make_new_table_all(self): # 모든 변동 주문을 반영한 생산계획 생성
        cobine_table = pd.merge(self.order_table, self.order_change_table, on='주문번호', suffixes=('_1', '_2'))
        cobine_table['주문량차이'] = cobine_table['주문량_2'] - cobine_table['주문량_1']
        cobine_table['납기일차이'] = (pd.to_datetime(cobine_table['납기일_2']) - pd.to_datetime(cobine_table['납기일_1'])).dt.days
        differ_table = cobine_table.loc[(cobine_table['주문량차이'] != 0) | (cobine_table['납기일차이'] != 0)]
        result_table = differ_table[['주문번호', '제품명_1', '제품번호_1', '주문량차이', '납기일차이']]
        for order_number, _ in result_table.iterrows():
            self.make_new_table(order_number)
        original_table = load_workbook(filename='생산_시간별_시퀀스.xlsx')
        new_table = load_workbook(filename='new_생산계획_test.xlsx')
        yellow_fill = PatternFill(start_color="FFFF09", end_color="FFFF09", fill_type="solid")
        sheet1 = original_table.active
        sheet2 = new_table.active

        for row in range(1, sheet2.max_row + 1):
            for col in range(2, sheet2.max_column + 1):
                cell1 = sheet1.cell(row=row, column=col)
                cell2 = sheet2.cell(row=row, column=col)

                if cell1.value != cell2.value:
                    cell2.fill = yellow_fill
        new_table.save(filename='new_생산계획_test.xlsx')
        result_profit = self.all_profit()
        print(result_profit)
        return result_profit

    def time_check(self): # 해당 주문이 납기일을 지키는지 판단
        for order_number in range(len(self.order_table)):
            target_order = self.order_change_table[self.order_change_table['주문번호'] == order_number]
            target_num = target_order.iloc[0,2]
            due_date = pd.to_datetime(target_order.iloc[0, 4]) # 납기일
            due_date = due_date.replace(second=0, microsecond=0)
            end_date = pd.to_datetime(self.result.loc[self.order_change_table['주문번호'] == order_number, '공정마감일'].iloc[0])
            end_date = end_date.replace(second=0, microsecond=0) # 마감일
            end_idx = self.result.loc[self.order_change_table['주문번호'] == order_number, '공정마감인덱스'].iloc[0]
            production_profit = self.result.loc[self.result['주문번호'] == order_number, '주문수익'].iloc[0]  # 주문 수익
            target_production = target_order.iloc[0, 3]  # 주문량
            result = self.result.loc[self.result['주문번호'] == order_number, '생산량'].iloc[0] # 생산량

            if end_date == due_date:  # 납기일과 마감일 동일
                self.result.loc[self.result['주문번호'] == order_number, '주문수익'] = production_profit
            elif end_date < due_date and end_idx != 0:  # 납기일 이내 마감 했을때
                gap = ((due_date - end_date).total_seconds() / 3600) // 24
                self.result.loc[self.result['주문번호'] == order_number, '주문수익'] = production_profit - gap * self.cost_table.loc[
                    target_num, '재고보유비용']
            elif end_date > due_date and end_date <= due_date + pd.Timedelta(days=10): # 납기일을 못지켰지만 납기일 +10일까지 마감일이 들어올 때
                gap = ((end_date - due_date).total_seconds() / 3600) // 24
                self.result.loc[self.result['주문번호'] == order_number, '주문수익'] = production_profit - gap * (
                self.cost_table.loc[target_num, '납기 지연 패널티 비용'])
            elif end_date > due_date + pd.Timedelta(days=10): # 마감일이 납기일 10일을 넘기는 경우
                gap = ((end_date - due_date).total_seconds() / 3600) // 24
                self.result.loc[self.result['주문번호'] == order_number, '주문수익'] = production_profit - (target_production - result) * \
                                                                    self.cost_table.loc[target_num, '납품 미달 패널티 비용'] - gap * (
                                                                    self.cost_table.loc[target_num, '납기 지연 패널티 비용'])
            elif end_idx == 0: #전체 납품 미달 패널티
                self.result.loc[self.result['주문번호'] == order_number, '주문수익'] = 0 - target_production * \
                                                                    self.cost_table.loc[target_num, '납품 미달 패널티 비용']

    def get_production_quantity(self, order_number): # 해당 주문의 생산량 계산
        target_pro = self.order_change_table[self.order_change_table['주문번호'] == order_number].iloc[0,1]
        idx = self.make_blk(order_number, self.production_plan)
        production_quantity = 0
        for i in range(idx[0], idx[1] + 1):
            if self.production_plan.loc[i, target_pro] != 's':
                production_quantity += self.production_plan.loc[i, target_pro]
        return production_quantity

    def table_check(self): # 생산 계획 테이블을 기반으로 생산량, 공정 시작일, 마감일, 셋업량, 공정시간, 패널티를 반영한 수익 계산
        for i in range(0,14): #주문 번호에 따라 result table 저장
            prod_order = self.order_change_table[self.order_change_table['주문번호'] == i] # 주문 정보
            prod_num = prod_order.iloc[0,2] # 제품 번호
            prod_name = prod_order.iloc[0,1] # 제품명
            result = self.get_production_quantity(i) # 생산량
            numb = self.make_blk(i, self.production_plan) # 공정 인덱스: first_numb, last_numb, blk_list
            end_date = self.production_plan.loc[numb[1], 'Date'] # 공정 마감일
            start_date = self.production_plan.loc[numb[0], 'Date'] #공정 시작일
            setup_counts = self.production_plan.loc[numb[0]:numb[1]+1, prod_name].apply(lambda x: x == 's').sum().sum() # 셋업량
            self.result.loc[self.result['주문번호'] == i, '생산량'] = result
            self.result.loc[self.result['주문번호'] == i, '공정마감일'] = end_date
            self.result.loc[self.result['주문번호'] == i, '공정시작일'] = start_date
            self.result.loc[self.result['주문번호'] == i, '공정마감인덱스'] = numb[1]
            self.result.loc[self.result['주문번호'] == i, '공정시작인덱스'] = numb[0]
            self.result.loc[self.result['주문번호'] == i, '셋업량'] = setup_counts
            self.result.loc[self.result['주문번호'] == i, '공정시간'] = self.result.loc[self.result['주문번호'] == i, '셋업량'].iloc[
                                                                    0] + \
                                                                self.result.loc[self.result['주문번호'] == i, '생산량'].iloc[
                                                                    0] * self.cost_table.loc[prod_num, '생산시간']
            self.result.loc[self.result['주문번호'] == i, '주문수익'] = result * (
                        self.cost_table.loc[prod_num, '이익'] - self.cost_table.loc[prod_num, '생산비용']) - setup_counts * \
                                                                self.cost_table.loc[prod_num, '셋업비용']
        return
  
    def profit(self, order_number):
        prod_profit = self.result.loc[self.result['주문번호'] == order_number, '주문수익'] 
        return prod_profit

    def all_profit(self): #전체 수익 계산
        real_profit = sum(self.result['주문수익'])
        return real_profit


Heuristic = Heuristic_Production_Plan('생산_시간별_시퀀스.xlsx', '비용_테이블.xlsx', '주문_테이블.xlsx', '주문_변동_테이블_6_4.xlsx')
start_time = time.time()
Heuristic.make_new_table_all()
print(time.time() - start_time)
